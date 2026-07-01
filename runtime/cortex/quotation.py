"""Quotation builder -> branded house-format PDF, delivered to the Inbox + R2 (see docs/COMPANY-STANDARD.md).

Mirrors seo_report.py: a company task asks for a quote, this renders a finished PDF and returns the pieces
to fill an Inbox card; engine.deliver_quotation drops the card and R2 stores the delivery copy.

Separation of concerns (logic-lives-in-skills):
  * This module is PLUMBING: it lays out whatever structured config it is handed and STAMPS the maths.
  * The WHAT — line-item skeletons per service type + the terms text — lives in editable DATA
    (`settings['quotation_presets']`), with DEFAULT_PRESETS here only as a seed/fallback.
  * House details (legal entity, bank, VAT, address, contact, brand) are read LIVE from the company's
    `company_profiles.data`, never hardcoded or invented.

No-invented-facts rule: prices come from the request (the operator states them). This module never makes a
price up; a line with no unit price renders blank and is counted in `blanks` so the card can flag it. All
totals (subtotal / agency fee / VAT / grand total) are computed here in code, never by the model.
"""
from __future__ import annotations

import base64
import datetime
import io
import os
import re

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (HRFlowable, Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table,
                                TableStyle)

from . import db, store

# ---------------------------------------------------------------------------
# Presets — the editable source is settings['quotation_presets']; this is the seed/fallback only.
# A preset = {title, agency_fee, note, sections:[{header, items:[str]}], terms:{intro, groups:[{heading,lines}]}}.
# Items are line-item LABELS only (no prices — prices are supplied per request).
# ---------------------------------------------------------------------------

_AI_TERMS = {
    "intro": "Sensa Productions is the trading brand of Sky Vision Aerial Photography Services, Dubai, UAE. "
             "These terms apply to the AI video production services set out in this quotation (the Customer "
             "being the client named above).",
    "groups": [
        {"heading": "Payment", "lines": [
            "1.  Down payment: A 70% down payment is required to commence the project. Work begins once it is received.",
            "2.  Balance: The remaining 30% is due before final delivery, on approval of the work (see Revisions & Delivery below).",
            "3.  Payments & VAT: All payments are made to Sky Vision Aerial Photography Services at the account on this quotation. Prices are in AED and exclusive of 5% VAT unless stated.",
        ]},
        {"heading": "Revisions & Delivery", "lines": [
            "4.  Revisions: Two revision requests are included, to be provided by the Customer in writing within two weeks of delivery. Requests made after two weeks, or beyond the two included rounds, are quoted separately.",
            "5.  Approval & watermarked file: Once the Customer is happy with the output, Sensa delivers a watermarked file for approval and requests final payment.",
            "6.  Final file: On receipt of final payment, Sensa releases the final unwatermarked file.",
            "7.  Final payment timing: Final payment is due upon approval and in any event before delivery of the unwatermarked file, and in any event within thirty days of the watermarked delivery.",
        ]},
        {"heading": "Ownership & AI", "lines": [
            "8.  Ownership: Until the balance is paid in full, all deliverables and related rights remain the property of Sensa Productions, which may withhold the unwatermarked file and recover any outstanding fees. Rights pass to the Customer on full payment.",
            "9.  AI output: Deliverables are produced with generative AI tools that Sensa directs and reviews. The final approved output is human reviewed; Sensa is not liable for outputs the Customer later edits or regenerates outside Sensa.",
            "10. Source & working files: Project and source files are not included and are charged separately if required. Sensa retains files only as needed to deliver the project.",
        ]},
        {"heading": "General", "lines": [
            "11. Scope: This quotation is based on the agreed concept and references. Material changes to the concept, scope, length or number of deliverables after sign-off will incur additional fees and may affect the schedule.",
            "12. Licensed music & assets: Any licensed music, stock or third-party assets are used under their own licences; extended or broadcast usage is quoted separately.",
            "13. Customer responsibilities: The Customer provides an accurate brief and timely approvals of key stages (for example the concept, script and key frames) so the schedule can be met.",
            "14. Confidentiality: Each party keeps the other's confidential information private.",
            "15. Liability: Sensa's total liability is limited to the fees paid under this quotation; Sensa is not liable for indirect or consequential loss.",
            "16. Cancellation: On cancellation the Customer is liable for all costs incurred to that point; the down payment is non-refundable once work has begun.",
            "17. Force majeure: Neither party is liable for delay or failure caused by events beyond its reasonable control.",
            "18. Governing law: These terms are governed by the laws of the Emirate of Dubai and the United Arab Emirates.",
        ]},
    ],
}

DEFAULT_PRESETS = {
    "ai-production": {
        "title": "AI VIDEO PRODUCTION QUOTATION",
        "agency_fee": False,
        "note": "Prices are added per the agreed scope; subtotal, VAT and total calculate automatically.",
        # What the client actually receives — shown in a Deliverables block at the top (editable data).
        "deliverables": [
            "1 x AI-generated video, up to 30 seconds",
            "Master file in 4K (3840 x 2160), MP4 (H.264)",
            "Aspect versions: 16:9 landscape, 9:16 vertical, 1:1 square",
            "Licensed music track (single-project licence)",
            "Delivered as a download link",
        ],
        # Each item is {desc, weight}: `weight` is the relative share used to split a stated total into
        # "fair rates" (editable data — tune these to reshape the breakdown, no code change).
        "sections": [
            {"header": "A ·  PRE-PRODUCTION", "items": [
                {"desc": "Concept development & creative direction", "weight": 3},
                {"desc": "Treatment & moodboard", "weight": 1.5},
                {"desc": "Scripting & narration", "weight": 1.5},
                {"desc": "Storyboard & shot planning", "weight": 2},
                {"desc": "Start-frame & end-frame design (key frames)", "weight": 2.5},
            ]},
            {"header": "B ·  GENERATION", "items": [
                {"desc": "Image generation (key frames & assets)", "weight": 2},
                {"desc": "Video generation (AI clips)", "weight": 4},
                {"desc": "Upscaling & frame interpolation", "weight": 1.5},
            ]},
            {"header": "C ·  POST-PRODUCTION", "items": [
                {"desc": "Editing & assembly", "weight": 3},
                {"desc": "Sound design", "weight": 2},
                {"desc": "Licensed music", "weight": 1},
                {"desc": "Voiceover & narration mix", "weight": 1.5},
                {"desc": "Colour & final grade", "weight": 1.5},
                {"desc": "Rendering & final master delivery", "weight": 1},
            ]},
        ],
        "terms": _AI_TERMS,
    },
}


def presets() -> dict:
    """Editable presets from settings, falling back to the built-in seed."""
    return db.setting_get("quotation_presets") or DEFAULT_PRESETS


def available() -> dict:
    """{preset_key: title} for the tool/UI."""
    return {k: v.get("title", k) for k, v in presets().items()}


# ---------------------------------------------------------------------------
# House details (live from the company profile) + small helpers
# ---------------------------------------------------------------------------

def _profile(company_id: int) -> dict:
    r = db.one("select data from company_profiles where company_id = %s", (company_id,))
    return (r or {}).get("data") or {}


def _vat_rate(data: dict) -> float:
    m = re.search(r"(\d+(?:\.\d+)?)", str(data.get("vat") or "5"))
    return (float(m.group(1)) / 100.0) if m else 0.05


def _money(v: float, cur: str) -> str:
    return f"{cur} {v:,.2f}"


def _next_number() -> str:
    """Real, code-stamped quote number: SEN-YYYY-NNNN with a monotonic counter in settings."""
    n = int(db.setting_get("quotation_seq") or 0) + 1
    db.setting_set("quotation_seq", n)
    return f"SEN-{datetime.date.today().year}-{n:04d}"


# ---------------------------------------------------------------------------
# Resolve — the shared model both renderers (PDF + XLSX) build from
# ---------------------------------------------------------------------------

def _resolve(company: str, preset: str, *, customer: str, sections, total, total_inclusive, title, note,
             agency_fee, terms, deliverables, number=None) -> dict:
    """Resolve house data + preset + priced line items into one model dict, stamping every figure in code.

    Pricing modes: `total` given -> "fair rates" split by per-item weight; `sections` with unit/qty ->
    explicit lines; neither -> blank skeleton. `total_inclusive` treats the stated figure as VAT-inclusive.
    """
    company = (company or "").lower()
    co = store.get_company_by_slug(company)
    if not co:
        raise ValueError(f"unknown company {company}")
    data = _profile(co["id"])
    pset = presets().get(preset) or DEFAULT_PRESETS["ai-production"]

    title = title or pset.get("title", "QUOTATION")
    note = note if note is not None else pset.get("note", "")
    agency_fee = pset.get("agency_fee", False) if agency_fee is None else agency_fee
    terms = terms or pset.get("terms") or {}
    deliverables = deliverables if deliverables is not None else list(pset.get("deliverables") or [])
    cur = (data.get("currency") or "AED").upper()
    vat_rate = _vat_rate(data)

    # Line items: from the request, else the preset skeleton (copy each item so we never mutate the preset).
    if sections is None:
        sections = [{"header": s["header"],
                     "items": [({**it} if isinstance(it, dict) else {"desc": it}) for it in s["items"]]}
                    for s in pset.get("sections", [])]

    stated = None
    blanks = 0
    if total is not None:
        stated = float(total)
        subtotal = round(stated / (1 + vat_rate), 2) if total_inclusive else stated
        flat = [it for s in sections for it in s.get("items", [])]
        wsum = sum(float(it.get("weight") or 1) for it in flat) or 1.0
        for it in flat:
            it["_amount"] = round(subtotal * float(it.get("weight") or 1) / wsum, 2)
        drift = round(subtotal - sum(it["_amount"] for it in flat), 2)
        if flat and drift:   # park the rounding remainder on the largest line so the split sums exactly
            big = max(flat, key=lambda it: it["_amount"])
            big["_amount"] = round(big["_amount"] + drift, 2)
        fee = 0.0
        if note == pset.get("note"):
            note = "Fees shown are a breakdown of the agreed project total; VAT calculates on top."
    else:
        subtotal = 0.0
        for s in sections:
            for it in s.get("items", []):
                unit = it.get("unit")
                qty = it.get("qty") or 1
                if unit in (None, ""):
                    blanks += 1
                    it["_amount"] = None
                else:
                    it["_amount"] = float(unit) * float(qty)
                    subtotal += it["_amount"]
        fee = round(subtotal * 0.15, 2) if agency_fee else 0.0
    vat = round((subtotal + fee) * vat_rate, 2)
    grand = round(subtotal + fee + vat, 2)
    number = number or _next_number()   # reuse a pinned number when rendering both formats of one quote
    n_items = len([1 for s in sections for _ in s.get("items", [])])
    blank_note = f" {blanks} price(s) left blank." if blanks else ""
    summary = (f"{co['name']} {preset} quotation {number}" + (f" for {customer}" if customer else "")
               + f": {n_items} line items, total {_money(grand, cur)} incl. VAT.{blank_note}")
    return {"company": company, "co": co, "data": data, "preset": preset, "title": title, "note": note,
            "agency_fee": agency_fee, "terms": terms, "deliverables": deliverables, "customer": customer,
            "cur": cur, "vat_rate": vat_rate, "sections": sections, "subtotal": subtotal, "fee": fee,
            "vat": vat, "grand": grand, "number": number, "stated": stated, "blanks": blanks,
            "summary": summary}


def _return(m: dict, path: str) -> dict:
    return {"path": path, "number": m["number"], "title": m["title"], "summary": m["summary"],
            "company": m["company"], "customer": m["customer"], "total": m["grand"], "currency": m["cur"],
            "blanks": m["blanks"], "stated": m["stated"], "preset": m["preset"]}


# ---------------------------------------------------------------------------
# Render — PDF
# ---------------------------------------------------------------------------

def generate(company: str, preset: str = "ai-production", *, customer: str = "", sections: list | None = None,
             total: float | None = None, total_inclusive: bool = False, title: str | None = None,
             note: str | None = None, agency_fee: bool | None = None, terms: dict | None = None,
             deliverables: list | None = None, number: str | None = None, out_dir: str = "/tmp") -> dict:
    """Build the house-format quotation PDF for `company`. See `_resolve` for the pricing modes.
    Returns {path, number, title, summary, company, customer, total, currency, blanks, stated, preset}.
    """
    m = _resolve(company, preset, customer=customer, sections=sections, total=total,
                 total_inclusive=total_inclusive, title=title, note=note, agency_fee=agency_fee,
                 terms=terms, deliverables=deliverables, number=number)
    company, co, data, title, note = m["company"], m["co"], m["data"], m["title"], m["note"]
    agency_fee, terms, deliverables = m["agency_fee"], m["terms"], m["deliverables"]
    cur, vat_rate, sections = m["cur"], m["vat_rate"], m["sections"]
    subtotal, fee, vat, grand = m["subtotal"], m["fee"], m["vat"], m["grand"]
    number, blanks, stated, customer = m["number"], m["blanks"], m["stated"], m["customer"]
    today = datetime.date.today().strftime("%d %b %Y")

    # ---- brand ----
    brand = (data.get("brand") or {})
    palette = brand.get("colors") or {}
    BG = colors.HexColor(palette.get("bg", "#0A0A0A"))
    ACCENT = colors.HexColor(palette.get("primary", "#00DAFF"))
    INK = colors.HexColor("#15202b")
    GREY = colors.HexColor("#667")
    LIGHT = colors.HexColor("#eef3f5")
    WHITE = colors.white

    ss = getSampleStyleSheet()
    NAME = ParagraphStyle("NAME", parent=ss["Normal"], fontSize=15, textColor=WHITE, fontName="Helvetica-Bold",
                          leading=17)
    BANDSUB = ParagraphStyle("BANDSUB", parent=ss["Normal"], fontSize=9, textColor=WHITE, leading=12)
    H1 = ParagraphStyle("H1", parent=ss["Title"], fontSize=17, textColor=INK, spaceAfter=1, alignment=0)
    SUB = ParagraphStyle("SUB", parent=ss["Normal"], fontSize=9, textColor=GREY, leading=12)
    SECT = ParagraphStyle("SECT", parent=ss["Normal"], fontSize=9.5, textColor=colors.white,
                          fontName="Helvetica-Bold")
    TD = ParagraphStyle("TD", parent=ss["Normal"], fontSize=9, textColor=INK, leading=12)
    TDR = ParagraphStyle("TDR", parent=TD, alignment=2)
    TH = ParagraphStyle("TH", parent=ss["Normal"], fontSize=8, textColor=ACCENT, fontName="Helvetica-Bold")
    THR = ParagraphStyle("THR", parent=TH, alignment=2)
    CAP = ParagraphStyle("CAP", parent=ss["Normal"], fontSize=9.5, textColor=INK, fontName="Helvetica-Bold",
                         spaceBefore=9, spaceAfter=3)
    SMALL = ParagraphStyle("SMALL", parent=ss["Normal"], fontSize=8, textColor=GREY, leading=11)
    TERMH = ParagraphStyle("TERMH", parent=ss["Normal"], fontSize=10, textColor=ACCENT,
                           fontName="Helvetica-Bold", spaceBefore=8, spaceAfter=3)
    TERML = ParagraphStyle("TERML", parent=ss["Normal"], fontSize=8.2, textColor=INK, leading=11, spaceAfter=2)

    # ---- header band (logo on the brand background) ----
    logo_flow = Paragraph(co["name"], NAME)
    b64 = brand.get("logo_dark_b64")
    if b64:
        try:
            raw = base64.b64decode(b64.split(",")[-1])
            img = Image(io.BytesIO(raw))
            ratio = img.imageHeight / float(img.imageWidth or 1)
            img.drawWidth = 42 * mm
            img.drawHeight = min(14 * mm, 42 * mm * ratio)
            logo_flow = img
        except Exception:  # noqa: BLE001 — bad/absent logo must never break a quote
            pass
    band_right = Paragraph("QUOTATION", ParagraphStyle("QN", parent=NAME, alignment=2, fontSize=16))
    band = Table([[logo_flow, band_right]], colWidths=[100 * mm, 74 * mm],
                 style=TableStyle([("BACKGROUND", (0, 0), (-1, -1), BG),
                                   ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                                   ("LEFTPADDING", (0, 0), (0, 0), 10), ("RIGHTPADDING", (-1, 0), (-1, 0), 10),
                                   ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 8)]))

    # ---- meta (title + quote details / customer) ----
    meta_left = [Paragraph(title, H1), Paragraph(f"Quotation {number} &nbsp;·&nbsp; {today}", SUB)]
    meta_right = [Paragraph("PREPARED FOR", THR),
                  Paragraph(customer or "&nbsp;", TDR),
                  Spacer(1, 4),
                  Paragraph("VALID FOR 30 DAYS", THR)]
    meta = Table([[meta_left, meta_right]], colWidths=[110 * mm, 64 * mm],
                 style=TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))

    story: list = [band, Spacer(1, 8), meta,
                   HRFlowable(width="100%", thickness=1, color=ACCENT, spaceBefore=8, spaceAfter=8)]

    # ---- deliverables (what the client receives) ----
    if deliverables:
        story.append(Paragraph("DELIVERABLES", TH))
        for d in deliverables:
            story.append(Paragraph(f"•&nbsp;&nbsp;{d}", TD))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#dde"),
                                spaceBefore=7, spaceAfter=8))

    # ---- line-item table ----
    rows = [[Paragraph("DESCRIPTION", TH), Paragraph("QTY", THR), Paragraph("UNIT", THR),
             Paragraph("AMOUNT", THR)]]
    tstyle = [("LINEBELOW", (0, 0), (-1, 0), 0.8, ACCENT),
              ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
              ("VALIGN", (0, 0), (-1, -1), "MIDDLE")]
    r = 1
    for s in sections:
        rows.append([Paragraph(s["header"], SECT), "", "", ""])
        tstyle += [("BACKGROUND", (0, r), (-1, r), INK), ("SPAN", (0, r), (-1, r)),
                   ("TOPPADDING", (0, r), (-1, r), 4), ("BOTTOMPADDING", (0, r), (-1, r), 4)]
        r += 1
        for it in s.get("items", []):
            unit = it.get("unit")
            qty = it.get("qty") or 1
            amt = it.get("_amount")
            rows.append([Paragraph(it["desc"], TD),
                         Paragraph(str(qty) if unit not in (None, "") else "", TDR),
                         Paragraph(_money(float(unit), cur) if unit not in (None, "") else "", TDR),
                         Paragraph(_money(amt, cur) if amt is not None else "", TDR)])
            tstyle.append(("LINEBELOW", (0, r), (-1, r), 0.25, colors.HexColor("#dde")))
            r += 1
    items_tbl = Table(rows, colWidths=[104 * mm, 16 * mm, 27 * mm, 27 * mm], style=TableStyle(tstyle),
                      repeatRows=1)
    story.append(items_tbl)

    # ---- totals ----
    tot_rows = [["", Paragraph("Subtotal", TD), Paragraph(_money(subtotal, cur), TDR)]]
    if agency_fee:
        tot_rows.append(["", Paragraph("Agency fee (15%)", TD), Paragraph(_money(fee, cur), TDR)])
    tot_rows.append(["", Paragraph(f"VAT ({int(vat_rate*100)}%)", TD), Paragraph(_money(vat, cur), TDR)])
    tot_rows.append(["", Paragraph("<b>Total</b>", TD), Paragraph(f"<b>{_money(grand, cur)}</b>", TDR)])
    tr = len(tot_rows) - 1
    totals = Table(tot_rows, colWidths=[104 * mm, 43 * mm, 27 * mm],
                   style=TableStyle([("LINEABOVE", (1, tr), (-1, tr), 0.8, ACCENT),
                                     ("TOPPADDING", (0, 0), (-1, -1), 2), ("BOTTOMPADDING", (0, 0), (-1, -1), 2)]))
    story += [Spacer(1, 2), totals]
    if note:
        story.append(Paragraph(f"<i>{note}</i>", SMALL))

    # ---- payment + bank band ----
    bank = (data.get("bank_details") or "").replace("\n", "<br/>")
    pay_left = [Paragraph("PAYMENT", TH),
                Paragraph("70% down payment to commence; 30% balance before final delivery, on approval.", TD),
                Paragraph(f"All prices in {cur}, exclusive of {int(vat_rate*100)}% VAT.", SMALL)]
    pay_right = [Paragraph("BANK DETAILS", TH), Paragraph(bank or "&nbsp;", SMALL)]
    pay = Table([[pay_left, pay_right]], colWidths=[100 * mm, 74 * mm],
                style=TableStyle([("BACKGROUND", (0, 0), (-1, -1), LIGHT), ("VALIGN", (0, 0), (-1, -1), "TOP"),
                                  ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                                  ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 8)]))
    story += [Spacer(1, 10), pay]

    # ---- acceptance ----
    acc = Table([[Paragraph("Accepted by (name & signature)", SMALL), Paragraph("Date", SMALL)],
                 [Paragraph("&nbsp;", TD), Paragraph("&nbsp;", TD)]],
                colWidths=[120 * mm, 54 * mm],
                style=TableStyle([("LINEBELOW", (0, 1), (0, 1), 0.5, GREY), ("LINEBELOW", (1, 1), (1, 1), 0.5, GREY),
                                  ("TOPPADDING", (0, 1), (-1, 1), 16)]))
    story += [Spacer(1, 12), acc]

    # ---- page 2: terms ----
    if terms:
        story.append(PageBreak())
        story.append(Paragraph("Terms &amp; Conditions", H1))
        if terms.get("intro"):
            story.append(Paragraph(terms["intro"], ParagraphStyle("TI", parent=SMALL, spaceAfter=6)))
        story.append(HRFlowable(width="100%", thickness=1, color=ACCENT, spaceBefore=4, spaceAfter=6))
        for g in terms.get("groups", []):
            story.append(Paragraph(g["heading"], TERMH))
            for ln in g.get("lines", []):
                story.append(Paragraph(ln, TERML))

    # ---- footer on every page (compact: brand + contact; the legal entity sits in the terms + bank) ----
    phone = (data.get("phone") or "").split(";")[0].strip()
    email = data.get("inbox_email") or ""
    foot = " · ".join(x for x in [co["name"], phone, email] if x)

    def _footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(GREY)
        canvas.drawString(18 * mm, 8 * mm, foot[:120])
        canvas.drawRightString(A4[0] - 18 * mm, 8 * mm, f"Page {doc_.page}")
        canvas.restoreState()

    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, f"quotation-{company}-{number}.pdf")
    SimpleDocTemplate(path, pagesize=A4, topMargin=14 * mm, bottomMargin=16 * mm, leftMargin=18 * mm,
                      rightMargin=18 * mm, title=f"{co['name']} Quotation {number}").build(
        story, onFirstPage=_footer, onLaterPages=_footer)

    blank_note = f" {blanks} price(s) left blank." if blanks else ""
    n_items = len([1 for s in sections for _ in s.get("items", [])])
    summary = (f"{co['name']} {preset} quotation {number}"
               + (f" for {customer}" if customer else "")
               + f": {n_items} line items, total {_money(grand, cur)} incl. VAT.{blank_note}")
    return {"path": path, "number": number, "title": title, "summary": summary, "company": company,
            "customer": customer, "total": grand, "currency": cur, "blanks": blanks, "stated": stated,
            "preset": preset}


# ---------------------------------------------------------------------------
# Render — XLSX (same house format, editable, with live-recalculating totals)
# ---------------------------------------------------------------------------

def generate_xlsx(company: str, preset: str = "ai-production", *, customer: str = "",
                  sections: list | None = None, total: float | None = None, total_inclusive: bool = False,
                  title: str | None = None, note: str | None = None, agency_fee: bool | None = None,
                  terms: dict | None = None, deliverables: list | None = None, number: str | None = None,
                  out_dir: str = "/tmp") -> dict:
    """Build the house-format quotation as an editable .xlsx: brand band, a Deliverables block, the line-item
    table, then Subtotal/VAT/Total as live formulas (edit a line amount and it recalculates), payment + bank
    band, acceptance, and Terms on a second sheet. Same model as the PDF (see `_resolve`)."""
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    m = _resolve(company, preset, customer=customer, sections=sections, total=total,
                 total_inclusive=total_inclusive, title=title, note=note, agency_fee=agency_fee,
                 terms=terms, deliverables=deliverables, number=number)
    co, data, cur, vat_rate = m["co"], m["data"], m["cur"], m["vat_rate"]
    brand = data.get("brand") or {}
    palette = brand.get("colors") or {}
    BG = palette.get("bg", "#0A0A0A").lstrip("#")
    ACCENT = palette.get("primary", "#00DAFF").lstrip("#")
    INK, GREY, LIGHT, DDE, WHITE = "15202B", "667788", "EEF3F5", "DDDDEE", "FFFFFF"
    money = f'"{cur}" #,##0.00'
    L = Alignment(horizontal="left", vertical="center", wrap_text=True)
    Rt = Alignment(horizontal="right", vertical="center")
    C = Alignment(horizontal="center", vertical="center")
    fillOf = lambda h: PatternFill("solid", fgColor=h)  # noqa: E731
    accent_b = Border(bottom=Side(style="medium", color=ACCENT))
    dde_b = Border(bottom=Side(style="thin", color=DDE))
    grey_b = Border(bottom=Side(style="thin", color="AAB1BA"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    ws.sheet_view.showGridLines = False
    for c, w in {"A": 56, "B": 8, "C": 17, "D": 17}.items():
        ws.column_dimensions[c].width = w

    # --- header band (rows 1-2): dark fill, logo left, QUOTATION right ---
    for rr in (1, 2):
        ws.row_dimensions[rr].height = 22
        for cc in "ABCD":
            ws[f"{cc}{rr}"].fill = fillOf(BG)
    ws["D2"] = "QUOTATION"; ws["D2"].font = Font(bold=True, size=15, color=WHITE); ws["D2"].alignment = Rt
    b64 = brand.get("logo_dark_b64")
    if b64:
        try:
            xi = XLImage(io.BytesIO(base64.b64decode(b64.split(",")[-1])))
            ratio = xi.width / float(xi.height or 1)
            xi.height = 30; xi.width = 30 * ratio
            xi.anchor = "A1"
            ws.add_image(xi)
        except Exception:  # noqa: BLE001 — bad/absent logo must never break the sheet
            pass

    r = 4
    ws[f"A{r}"] = m["title"]; ws[f"A{r}"].font = Font(bold=True, size=15, color=INK)
    ws[f"D{r}"] = "PREPARED FOR"; ws[f"D{r}"].font = Font(bold=True, size=8, color=ACCENT); ws[f"D{r}"].alignment = Rt
    ws[f"A{r+1}"] = f"Quotation {m['number']}  ·  {datetime.date.today().strftime('%d %b %Y')}"
    ws[f"A{r+1}"].font = Font(size=9, color=GREY)
    ws[f"D{r+1}"] = m["customer"] or ""; ws[f"D{r+1}"].alignment = Rt; ws[f"D{r+1}"].font = Font(size=10, color=INK)
    ws[f"D{r+2}"] = "VALID FOR 30 DAYS"; ws[f"D{r+2}"].font = Font(bold=True, size=8, color=ACCENT); ws[f"D{r+2}"].alignment = Rt
    r += 3

    # --- deliverables ---
    if m["deliverables"]:
        r += 1
        ws[f"A{r}"] = "DELIVERABLES"; ws[f"A{r}"].font = Font(bold=True, size=9, color=ACCENT)
        r += 1
        for d in m["deliverables"]:
            ws.merge_cells(f"A{r}:D{r}")
            ws[f"A{r}"] = f"•  {d}"; ws[f"A{r}"].font = Font(size=10, color=INK); ws[f"A{r}"].alignment = L
            r += 1
        for cc in "ABCD":
            ws[f"{cc}{r}"].border = dde_b
        r += 1

    # --- line-item table header ---
    hdr = {"A": "DESCRIPTION", "B": "QTY", "C": "UNIT", "D": "AMOUNT"}
    for cc, lab in hdr.items():
        cell = ws[f"{cc}{r}"]; cell.value = lab; cell.font = Font(bold=True, size=8, color=ACCENT)
        cell.border = accent_b; cell.alignment = Rt if cc in "BCD" else L
    r += 1
    amt_first = None
    for s in m["sections"]:
        ws.merge_cells(f"A{r}:D{r}")
        ws[f"A{r}"] = s["header"]; ws[f"A{r}"].font = Font(bold=True, size=9, color=WHITE)
        for cc in "ABCD":
            ws[f"{cc}{r}"].fill = fillOf(INK)
        r += 1
        for it in s.get("items", []):
            unit = it.get("unit"); qty = it.get("qty") or 1; amt = it.get("_amount")
            ws[f"A{r}"] = it["desc"]; ws[f"A{r}"].font = Font(size=10, color=INK); ws[f"A{r}"].alignment = L
            if unit not in (None, ""):        # explicit unit pricing -> editable qty x unit formula
                ws[f"B{r}"] = qty; ws[f"B{r}"].alignment = C
                ws[f"C{r}"] = float(unit); ws[f"C{r}"].number_format = money; ws[f"C{r}"].alignment = Rt
                ws[f"D{r}"] = f"=B{r}*C{r}"
            elif amt is not None:             # allocation ("fair rates") -> fixed line amount
                ws[f"D{r}"] = amt
            else:                             # blank template -> editable formula (0 until qty+unit typed)
                ws[f"C{r}"].number_format = money; ws[f"C{r}"].alignment = Rt; ws[f"B{r}"].alignment = C
                ws[f"D{r}"] = f"=B{r}*C{r}"
            ws[f"D{r}"].number_format = money; ws[f"D{r}"].alignment = Rt
            for cc in "ABCD":
                ws[f"{cc}{r}"].border = dde_b
            amt_first = amt_first or r
            r += 1
    amt_last = r - 1

    # --- totals (live formulas over the amount column) ---
    rng = f"D{amt_first}:D{amt_last}" if amt_first else "D1:D1"
    r += 1
    ws[f"C{r}"] = "Subtotal"; ws[f"C{r}"].alignment = Rt; ws[f"C{r}"].font = Font(size=10, color=INK)
    ws[f"D{r}"] = f"=SUM({rng})"; ws[f"D{r}"].number_format = money; ws[f"D{r}"].alignment = Rt
    sub_row = r
    if m["agency_fee"]:
        r += 1
        ws[f"C{r}"] = "Agency fee (15%)"; ws[f"C{r}"].alignment = Rt; ws[f"C{r}"].font = Font(size=10, color=INK)
        ws[f"D{r}"] = f"=D{sub_row}*0.15"; ws[f"D{r}"].number_format = money; ws[f"D{r}"].alignment = Rt
        fee_row = r
        base = f"(D{sub_row}+D{fee_row})"
    else:
        base = f"D{sub_row}"
    r += 1
    ws[f"C{r}"] = f"VAT ({int(vat_rate*100)}%)"; ws[f"C{r}"].alignment = Rt; ws[f"C{r}"].font = Font(size=10, color=INK)
    ws[f"D{r}"] = f"={base}*{vat_rate}"; ws[f"D{r}"].number_format = money; ws[f"D{r}"].alignment = Rt
    vat_row = r
    r += 1
    ws[f"C{r}"] = "Total"; ws[f"C{r}"].alignment = Rt; ws[f"C{r}"].font = Font(bold=True, size=11, color=INK)
    ws[f"D{r}"] = f"={base}+D{vat_row}"; ws[f"D{r}"].number_format = money; ws[f"D{r}"].alignment = Rt
    ws[f"D{r}"].font = Font(bold=True, size=11, color=INK)
    ws[f"C{r}"].border = accent_b; ws[f"D{r}"].border = accent_b
    if m["note"]:
        r += 1
        ws.merge_cells(f"A{r}:D{r}")
        ws[f"A{r}"] = m["note"]; ws[f"A{r}"].font = Font(italic=True, size=8, color=GREY); ws[f"A{r}"].alignment = L

    # --- payment + bank band ---
    r += 2
    pay_top = r
    ws[f"A{r}"] = "PAYMENT"; ws[f"A{r}"].font = Font(bold=True, size=8, color=ACCENT)
    ws[f"C{r}"] = "BANK DETAILS"; ws[f"C{r}"].font = Font(bold=True, size=8, color=ACCENT)
    r += 1
    ws.merge_cells(f"A{r}:B{r+2}")
    ws[f"A{r}"] = ("70% down payment to commence; 30% balance before final delivery, on approval.\n"
                   f"All prices in {cur}, exclusive of {int(vat_rate*100)}% VAT.")
    ws[f"A{r}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws[f"A{r}"].font = Font(size=9, color=INK)
    ws.merge_cells(f"C{r}:D{r+2}")
    ws[f"C{r}"] = data.get("bank_details") or ""
    ws[f"C{r}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws[f"C{r}"].font = Font(size=8, color=INK)
    for rr in range(pay_top, r + 3):
        ws.row_dimensions[rr].height = 15
        for cc in "ABCD":
            ws[f"{cc}{rr}"].fill = fillOf(LIGHT)
    r += 3

    # --- acceptance ---
    r += 2
    ws[f"A{r}"] = "Accepted by (name & signature)"; ws[f"A{r}"].font = Font(size=8, color=GREY)
    ws[f"C{r}"] = "Date"; ws[f"C{r}"].font = Font(size=8, color=GREY)
    r += 2
    ws[f"A{r}"].border = grey_b; ws[f"B{r}"].border = grey_b
    ws[f"C{r}"].border = grey_b; ws[f"D{r}"].border = grey_b

    # --- footer line ---
    r += 2
    phone = (data.get("phone") or "").split(";")[0].strip()
    ws.merge_cells(f"A{r}:D{r}")
    ws[f"A{r}"] = " · ".join(x for x in [co["name"], phone, data.get("inbox_email") or ""] if x)
    ws[f"A{r}"].font = Font(size=7, color=GREY)

    # --- sheet 2: terms ---
    terms = m["terms"]
    if terms:
        t2 = wb.create_sheet("Terms & Conditions")
        t2.sheet_view.showGridLines = False
        t2.column_dimensions["A"].width = 118
        tr = 1
        t2[f"A{tr}"] = "Terms & Conditions"; t2[f"A{tr}"].font = Font(bold=True, size=14, color=INK); tr += 2
        if terms.get("intro"):
            t2[f"A{tr}"] = terms["intro"]; t2[f"A{tr}"].font = Font(size=9, color=INK)
            t2[f"A{tr}"].alignment = Alignment(wrap_text=True, vertical="top"); t2.row_dimensions[tr].height = 30
            tr += 1
        for g in terms.get("groups", []):
            tr += 1
            t2[f"A{tr}"] = g["heading"]; t2[f"A{tr}"].font = Font(bold=True, size=10, color=ACCENT); tr += 1
            for ln in g.get("lines", []):
                t2[f"A{tr}"] = ln; t2[f"A{tr}"].font = Font(size=9, color=INK)
                t2[f"A{tr}"].alignment = Alignment(wrap_text=True, vertical="top")
                t2.row_dimensions[tr].height = 26; tr += 1

    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, f"quotation-{m['company']}-{m['number']}.xlsx")
    wb.save(path)
    return _return(m, path)
